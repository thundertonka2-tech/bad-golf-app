#!/usr/bin/env python3
"""
Bad Golf — course mapping audit.  THREE-NINE AWARE.

Run:  python3 qa/bg_mapping_audit.py [-o BadGolf_Mapping_Audit_YYYY-MM-DD.xlsx]

Why this file exists
--------------------
Every mapping audit before 2026-08-26 was written ad hoc inside a chat session and
thrown away afterwards, so the same wrong assumption got re-derived every time:
that a course's scorecard lives in `shared:courses` and its rating lives in a
`tees[]` entry on that card.

For a wired three-nine facility BOTH are false.  Its pars, stroke indexes, tee
yardages and course ratings live in the `THREE_NINE_COURSES` constant inside
golf-app.html (and, for later additions, in `shared:multinine-configs`), stored
PER NINE, with the rating attached to the 18-hole COMBINATION -- because that is
how a 27-hole facility is actually rated.  Its GPS greens live under per-nine ids
(`base#nine`), never under the base id.

An audit that does not know this reports, every single run:
  * "no scorecard at all"        -- for facilities that are fully carded
  * "card but no rated tee"      -- for facilities that are fully rated
  * "not mapped / partially mapped" -- for facilities that are fully mapped
  * "GPS pars disagree with the scorecard" -- comparing two things that are both
    superseded by the config

On 2026-08-26 that was 15 courses flagged across items 11 and 12, every one of
them a false positive, plus a large share of items 7, 8 and 9.  Acting on the
report as written would have DESTROYED data: writing an 18-hole card into
`shared:courses` for a three-nine shadows the per-nine config.

The app itself has been right about this since v882 (`bgCourseStatus()` reads
`hasCard = !!parArr || isTN` and `hasRS = isTN || ...`).  Only the offline audit
was wrong.  Keep it that way: if you add a check here, ask first what a
three-nine facility looks like to it.
"""

import argparse, base64, datetime, json, math, os, re, sys, urllib.request, urllib.error
from collections import defaultdict

SUPABASE_URL = os.environ.get("BG_SUPABASE_URL", "https://ojclesuwxhtzvrymqrwg.supabase.co")
ANON_KEY = os.environ.get("BG_SUPABASE_ANON_KEY", "")
# The code-review queue is admin workflow data and is correctly blocked to anon by
# the v1229 RLS lockdown. Set this to populate the "Kevin Queue" tab; leave it unset
# and every other check still runs, that one tab just says why it is empty.
SERVICE_KEY = os.environ.get("BG_SUPABASE_SERVICE_KEY", "")
APP_HTML = os.environ.get("BG_APP_HTML", "golf-app.html")

# ----------------------------------------------------------------- fetch ----

def _get(path, params=None, tries=4, key=None):
    url = SUPABASE_URL.rstrip("/") + path
    if params:
        url += "?" + "&".join(f"{k}={v}" for k, v in params.items())
    key = key or ANON_KEY
    last = None
    for attempt in range(tries):
        req = urllib.request.Request(url, headers={
            "apikey": key, "Authorization": "Bearer " + key,
            "Accept": "application/json", "Accept-Encoding": "gzip",
        })
        try:
            with urllib.request.urlopen(req, timeout=180) as r:
                raw = r.read()
                if r.headers.get("Content-Encoding") == "gzip":
                    import gzip
                    raw = gzip.decompress(raw)
                return json.loads(raw.decode("utf-8"))
        except urllib.error.HTTPError as e:
            last = e
            if e.code in (500, 502, 503, 504, 520, 544):
                import time
                time.sleep(1.5 * (attempt + 1))
                continue
            raise
        except Exception as e:      # transient socket/timeout
            last = e
            import time
            time.sleep(1.5 * (attempt + 1))
    raise last

def singleton(code, key=None):
    rows = _get("/rest/v1/games", {"code": f"eq.{code}", "select": "data"}, key=key)
    return rows[0]["data"] if rows else None

def fetch_course_gps(page=150):
    """Page course_gps in small slices.

    `holes` carries the full per-hole geometry (tee/green polygons, centrelines),
    so a 1,000-row page is tens of megabytes and PostgREST answers 500. 150 is a
    size the gateway reliably serves; it halves automatically if a page still fails.
    """
    out, off = [], 0
    while True:
        try:
            rows = _get("/rest/v1/course_gps", {
                "select": "course_id,name,city,holes,source",
                "limit": page, "offset": off, "order": "course_id",
            })
        except Exception:
            if page <= 25:
                raise
            page //= 2
            sys.stderr.write(f"  course_gps page too large; retrying at {page}\n")
            continue
        out.extend(rows)
        sys.stderr.write(f"\r  course_gps {len(out)} rows")
        if len(rows) < page:
            sys.stderr.write("\n")
            return out
        off += len(rows)

# ------------------------------------------------- three-nine config load ----

def _match_brace(s, start):
    d = 0
    for i in range(start, len(s)):
        if s[i] == "{":
            d += 1
        elif s[i] == "}":
            d -= 1
            if d == 0:
                return i
    raise ValueError("unbalanced braces")

def load_three_nine_from_app(path=APP_HTML):
    """Parse THREE_NINE_COURSES out of golf-app.html.

    Returns {course_id: {"nines": {nine: {"pars": [...], "rated": bool}},
                         "rated": bool}}
    `rated` is true when the facility carries ANY course rating -- either
    per-nine `tees[].rating` or, far more commonly, a `combos` block giving the
    rating for each 18-hole pairing.  Both shapes count.  105 of the 110 wired
    facilities use the combos shape.
    """
    if not os.path.exists(path):
        sys.stderr.write(f"WARN: {path} not found; three-nine exemptions from code will be EMPTY\n")
        return {}
    src = open(path, encoding="utf-8", errors="replace").read()
    m = re.search(r"const\s+THREE_NINE_COURSES\s*=\s*\{", src)
    if not m:
        sys.stderr.write("WARN: THREE_NINE_COURSES not found in app html\n")
        return {}
    start = src.index("{", m.start())
    blob = src[start:_match_brace(src, start) + 1]

    out = {}
    for cid, const in re.findall(r"'([a-z0-9\-]+)'\s*:\s*([A-Z][A-Z0-9_]*)", blob):
        cm = re.search(r"(?:const|var|let)\s+" + const + r"\s*=\s*", src)
        if not cm:
            continue
        cs = src.index("{", cm.end())
        cfg_txt = src[cs:_match_brace(src, cs) + 1]
        cfg = _loose_js_object(cfg_txt)
        if cfg is None:
            continue
        out[cid] = _summarise_cfg(cfg)
    return out

def _loose_js_object(txt):
    """Parse a JS object literal that may use unquoted keys and trailing commas."""
    try:
        return json.loads(txt)
    except Exception:
        pass
    t = re.sub(r"//[^\n]*", "", txt)
    t = re.sub(r"/\*.*?\*/", "", t, flags=re.S)
    t = re.sub(r"([{,]\s*)([A-Za-z_$][\w$]*)\s*:", r'\1"\2":', t)
    t = re.sub(r"'([^'\\]*)'", lambda mm: json.dumps(mm.group(1)), t)
    t = re.sub(r",\s*([}\]])", r"\1", t)
    try:
        return json.loads(t)
    except Exception:
        return None

def _summarise_cfg(cfg):
    nines = {}
    for k, v in cfg.items():
        if k.startswith("_") or k in ("searchCombos", "combos"):
            continue
        if isinstance(v, dict) and isinstance(v.get("pars"), list):
            rated = any(
                isinstance(t, dict) and t.get("rating") is not None and t.get("slope") is not None
                for t in (v.get("tees") or [])
            )
            nines[k] = {"pars": v["pars"], "rated": rated}
    combo_rated = False
    combos = cfg.get("combos")
    if isinstance(combos, dict):
        for arr in combos.values():
            if isinstance(arr, list) and any(
                isinstance(t, dict) and t.get("rating") is not None for t in arr
            ):
                combo_rated = True
                break
    return {
        "nines": nines,
        "rated": combo_rated or any(n["rated"] for n in nines.values()),
        "source": "app",
    }

def load_three_nine_from_db():
    cfgs = singleton("shared:multinine-configs") or {}
    out = {}
    for cid, cfg in cfgs.items():
        if isinstance(cfg, dict):
            out[cid] = _summarise_cfg(cfg)
            out[cid]["source"] = "db"
    return out

# ------------------------------------------------------------- geometry ----

def km(lat1, lng1, lat2, lng2):
    if None in (lat1, lng1, lat2, lng2):
        return None
    r = math.radians
    a = math.sin(r(lat2 - lat1) / 2) ** 2 + math.cos(r(lat1)) * math.cos(r(lat2)) * math.sin(r(lng2 - lng1) / 2) ** 2
    return round(6371 * 2 * math.asin(min(1, math.sqrt(a))), 3)

def greens_of(holes):
    if not isinstance(holes, dict):
        return {}
    return {h: v for h, v in holes.items() if isinstance(v, dict) and v.get("mid")}

# ----------------------------------------------------------------- audit ----

def build(argv=None):
    ap = argparse.ArgumentParser(description="Bad Golf mapping audit (three-nine aware)")
    ap.add_argument("-o", "--out", default=None, help="xlsx output path")
    ap.add_argument("--app", default=APP_HTML, help="path to golf-app.html")
    ap.add_argument("--json", action="store_true", help="also dump findings as JSON")
    ap.add_argument("--prev", default=None,
                    help="a previous run's .json, to show movement in the Summary")
    ap.add_argument("--kevin-json", default=None,
                    help="pre-exported code-review queue rows, when the anon key cannot read it")
    args = ap.parse_args(argv)

    if not ANON_KEY:
        sys.exit("Set BG_SUPABASE_ANON_KEY (read-only anon key from golf-app.html).")

    today = datetime.date.today().isoformat()
    out_path = args.out or f"BadGolf_Mapping_Audit_{today}.xlsx"

    # ---- load -------------------------------------------------------------
    additions = singleton("shared:course-library-additions") or []
    deletes = set(singleton("shared:course-deletes") or [])
    verified = set(singleton("shared:course-verified") or [])
    gps_rows = fetch_course_gps()

    cards = {}
    base = singleton("shared:courses") or {}
    if isinstance(base, dict):
        cards.update(base)
    for i in range(64):                     # shards win over the base blob
        shard = singleton("shared:courses:%02x" % i)
        if isinstance(shard, dict):
            cards.update(shard)

    tn = load_three_nine_from_app(args.app)
    tn.update(load_three_nine_from_db())    # DB-wired facilities added later
    def is_tn(cid):
        return cid in tn

    lib = {}
    for e in additions:
        if isinstance(e, dict) and e.get("id") and e["id"] not in deletes:
            lib[e["id"]] = e

    gps = {r["course_id"]: r for r in gps_rows}
    # per-nine rollup: a three-nine's greens live under `base#nine`, never the base id
    per_nine = defaultdict(dict)
    for cid, row in gps.items():
        if "#" in cid:
            b, n = cid.split("#", 1)
            per_nine[b][n] = row

    findings = defaultdict(list)

    # ---- 1. no scorecard ---------------------------------------------------
    #   THREE-NINE EXEMPT: pars live in the config, per nine.
    for cid, e in lib.items():
        if is_tn(cid):
            continue
        c = cards.get(cid)
        pars = (c or {}).get("pars")
        if not isinstance(pars, list) or len(pars) not in (9, 18):
            findings["no_card"].append({
                "id": cid, "name": e.get("name"), "city": e.get("city"), "st": e.get("st"),
                "issue": "No scorecard at all",
            })

    # ---- 2. card but no rated tee -----------------------------------------
    #   THREE-NINE EXEMPT: ratings hang off the 18-hole combination in `combos`,
    #   not off a tees[] entry.  105 of 110 wired facilities use that shape.
    for cid, e in lib.items():
        if is_tn(cid):
            continue
        c = cards.get(cid)
        if not c:
            continue
        tees = c.get("tees") or []
        rated = any(
            isinstance(t, dict) and t.get("rating") is not None and t.get("slope") is not None
            for t in tees
        ) or (c.get("rating") is not None and c.get("slope") is not None)
        if not rated:
            findings["no_rating"].append({
                "id": cid, "name": e.get("name"), "city": e.get("city"), "st": e.get("st"),
                "issue": "Card but no rated tee",
            })

    # ---- 3. mapping coverage ----------------------------------------------
    #   THREE-NINE: expected = nines x 9, counted across the per-nine rows.
    #   Counting the base id here is what produced "Shangri La 0/18" for months.
    for cid, e in lib.items():
        if is_tn(cid):
            cfg = tn[cid]
            nines = cfg["nines"]
            exp = len(nines) * 9
            got = sum(len(greens_of((per_nine[cid].get(n) or {}).get("holes"))) for n in nines)
            missing = [n for n in nines
                       if len(greens_of((per_nine[cid].get(n) or {}).get("holes"))) < 9]
            legacy = len(greens_of((gps.get(cid) or {}).get("holes")))
            if got < exp:
                findings["coverage"].append({
                    "id": cid, "name": e.get("name"), "city": e.get("city"), "st": e.get("st"),
                    "greens": got, "expected": exp,
                    "issue": "Not mapped" if got == 0 else "Partially mapped",
                    "nines_missing": ", ".join(sorted(missing)),
                    "legacy_base_greens": legacy,
                    "note": ("%d greens sit under the BASE id where the app cannot see them - "
                             "try splitting them into nines by par match before re-mapping" % legacy)
                            if legacy else "",
                })
            continue
        row = gps.get(cid)
        got = len(greens_of((row or {}).get("holes")))
        c = cards.get(cid) or {}
        pars = c.get("pars") if isinstance(c.get("pars"), list) else None
        exp = e.get("holes") or (len(pars) if pars else 18)
        if got < exp:
            findings["coverage"].append({
                "id": cid, "name": e.get("name"), "city": e.get("city"), "st": e.get("st"),
                "greens": got, "expected": exp,
                "issue": "Not mapped" if got == 0 else "Partially mapped",
                "nines_missing": "", "legacy_base_greens": 0, "note": "",
            })

    # ---- 4. GPS pars vs scorecard pars ------------------------------------
    #   THREE-NINE: compare each nine's mapping against the CONFIG's pars.
    #   Comparing a base-id 18 against a DB card is comparing two artefacts the
    #   wiring already superseded, which is why 19 of the 44 hits on 2026-08-26
    #   were noise.
    for cid, e in lib.items():
        pairs = []
        if is_tn(cid):
            for n, meta in tn[cid]["nines"].items():
                row = per_nine[cid].get(n)
                g = greens_of((row or {}).get("holes"))
                for h, v in g.items():
                    try:
                        idx = int(h) - 1
                    except ValueError:
                        continue
                    if v.get("par") and 0 <= idx < len(meta["pars"]):
                        pairs.append((f"{n}:{h}", v["par"], meta["pars"][idx]))
        else:
            c = cards.get(cid) or {}
            pars = c.get("pars") if isinstance(c.get("pars"), list) else None
            if not pars:
                continue
            g = greens_of((gps.get(cid) or {}).get("holes"))
            for h, v in g.items():
                try:
                    idx = int(h) - 1
                except ValueError:
                    continue
                if v.get("par") and 0 <= idx < len(pars):
                    pairs.append((h, v["par"], pars[idx]))
        if len(pairs) >= 6:
            agree = sum(1 for _, a, b in pairs if a == b)
            if agree / len(pairs) < 0.70:
                bad = [f"h{h}: gps {a} vs card {b}" for h, a, b in pairs if a != b][:8]
                findings["par_mismatch"].append({
                    "id": cid, "name": e.get("name"), "city": e.get("city"), "st": e.get("st"),
                    "agree": f"{agree}/{len(pairs)}",
                    "source": (gps.get(cid) or {}).get("source", ""),
                    "verified": "YES" if cid in verified else "",
                    "detail": "; ".join(bad),
                    "reference": "three-nine config" if is_tn(cid) else "scorecard",
                })

    # ---- 5. shared GPS signature ------------------------------------------
    sig = defaultdict(list)
    for cid, row in gps.items():
        g = greens_of(row.get("holes"))
        if not g:
            continue
        key = "|".join(f"{h}:{round(v['mid'][1],5)},{round(v['mid'][0],5)}"
                       for h, v in sorted(g.items(), key=lambda kv: int(kv[0]) if kv[0].isdigit() else 0))
        sig[key].append(cid)
    for key, ids in sig.items():
        if len(ids) > 1:
            live = [i for i in ids if i in lib or "#" in i]
            # A three-nine's base row and its own base#nine rows are the SAME holes by
            # construction - the nine rows were split out of the base row. That is one
            # facility, not two courses sharing a mapping, so it is not a finding.
            # Same family of mistake as the three-nine blindness this script was written
            # to end: a checker that does not know about the nine wiring cries wolf.
            if len({i.split("#", 1)[0] for i in live}) < 2:
                continue
            if len(live) > 1:
                findings["shared_gps"].append({
                    "ids": ", ".join(sorted(live)),
                    "names": ", ".join(sorted((lib.get(i, {}) or {}).get("name", i) for i in live)),
                    "greens": len(greens_of(gps[live[0]]["holes"])),
                    "issue": "Same GPS mapping on more than one course",
                })

    # ---- 6. mapped far from the library pin -------------------------------
    for cid, e in lib.items():
        row = gps.get(cid)
        g = greens_of((row or {}).get("holes"))
        if not g or e.get("lat") is None:
            continue
        clat = sum(v["mid"][1] for v in g.values()) / len(g)
        clng = sum(v["mid"][0] for v in g.values()) / len(g)
        d = km(clat, clng, e["lat"], e["lng"])
        if d and d > 2:
            findings["wrong_course"].append({
                "id": cid, "name": e.get("name"), "city": e.get("city"), "st": e.get("st"),
                "km": d, "greens": len(g), "source": (row or {}).get("source", ""),
                "verified": "YES" if cid in verified else "",
                "note": "Manual map - suspect the library PIN before the map"
                        if (row or {}).get("source") == "manual" else
                        "Check whether the pin or the mapping is wrong",
            })

    # ---- 7. city hygiene ---------------------------------------------------
    for cid, e in lib.items():
        city = (e.get("city") or "").strip()
        if not city or city == (e.get("st") or "").strip():
            findings["city"].append({
                "id": cid, "name": e.get("name"), "st": e.get("st"),
                "lat": e.get("lat"), "lng": e.get("lng"),
                "issue": "City blank or just the state code",
            })

    # ---- 8. Kevin's open queue --------------------------------------------
    #   The only items that reach him are courses still LIVE in the library whose
    #   most recent code-review entry sent them back. A course that has since been
    #   deleted still carries its old entry; it is not work, so it is not listed.
    try:
        if args.kevin_json and os.path.exists(args.kevin_json):
            findings["kevin"].extend(json.load(open(args.kevin_json)))
            raise StopIteration
        crq = singleton("shared:code-review-queue", key=SERVICE_KEY or None)
        if crq is None:
            raise RuntimeError(
                "not readable with this key - the code-review queue is admin data and "
                "anon is blocked by design (v1229 RLS). Set BG_SUPABASE_SERVICE_KEY to fill this tab.")
        latest = {}
        for e in crq:
            if not isinstance(e, dict):
                continue
            cid = ((e.get("course") or {}).get("id"))
            if not cid:
                continue
            stamp = str(e.get("lastEventAt") or e.get("resolvedAt") or e.get("rejectedAt") or e.get("createdAt") or "")
            if cid not in latest or stamp >= latest[cid][0]:
                latest[cid] = (stamp, e)
        for cid, (_stamp, e) in sorted(latest.items()):
            if e.get("status") != "rejected" or cid not in lib:
                continue
            d = (e.get("directive") or "").strip()
            if not d:
                note = str(e.get("resolutionNote") or "")
                d = note.split("\n\n")[0].strip()[:240]
            findings["kevin"].append({
                "id": cid, "name": lib[cid].get("name"), "city": lib[cid].get("city"),
                "st": lib[cid].get("st"), "status": "sent back", "directive": d,
            })
    except StopIteration:
        pass
    except Exception as exc:
        sys.stderr.write("NOTE: Kevin Queue tab left empty - %s\n" % exc)
        findings["kevin_note"] = str(exc)

    # ---- metrics: the size of the work, not the number of rows it lands on --
    nines_missing = greens_missing = 0
    vshort = vshort_greens = 0
    recovered = sum(1 for r in gps_rows
                    if str(r.get("source") or "").startswith("ninesplit-")) * 9
    for cid in lib:
        if cid not in tn:
            continue
        nines = tn[cid]["nines"]
        short = [n for n in nines
                 if len(greens_of((per_nine.get(cid, {}).get(n) or {}).get("holes"))) < 9]
        if not short:
            continue
        got = sum(len(greens_of((per_nine.get(cid, {}).get(n) or {}).get("holes")))
                  for n in nines)
        if cid in verified:
            vshort += 1
            vshort_greens += len(nines) * 9 - got
            continue
        nines_missing += len(short)
        greens_missing += len(nines) * 9 - got

    summary = {
        "generated": today,
        "live_courses": len(lib),
        "gps_rows": len(gps),
        "three_nine_facilities": len(tn),
        "counts": {k: len(v) for k, v in findings.items() if isinstance(v, list)},
        "metrics": {
            "nines_missing": nines_missing,
            "greens_missing": greens_missing,
            "greens_recovered": recovered,
            "verified_short": vshort,
            "verified_short_greens": vshort_greens,
            "reconciled": len(findings.get("kevin", [])),
        },
    }
    return summary, findings, tn, out_path, args

# ------------------------------------------------------------------ xlsx ----

# Severity is fixed per finding so the Summary reads the same way run to run --
# that is what makes the count column a progress bar rather than a snapshot.
SEVERITY = {
    "no_card": "LOW", "no_rating": "LOW", "coverage": "MEDIUM",
    "par_mismatch": "MEDIUM", "shared_gps": "HIGH", "wrong_course": "HIGH",
    "city": "LOW", "kevin": "INFO",
}
NOTES = {
    "no_card": "Wired three-nines are exempt - their pars live in the app config.",
    "no_rating": "Wired three-nines are exempt - ratings hang off the 18-hole combination.",
    "coverage": "Three-nine facilities are counted across per-nine rows (base#nine). "
                "'legacy base greens' means mapping exists under the base id where the app cannot see it.",
    "par_mismatch": "Three-nines are compared against the config's per-nine pars, not the DB card.",
    "shared_gps": "Two courses on one mapping. Players on one get the other's distances.",
    "wrong_course": "Green centroid >2 km from the library pin. On a manual map, suspect the PIN first.",
    "city": "Cosmetic, but it is what breaks Kevin's search.",
    "kevin": "Open items in the code-review queue, with the directive each one carries.",
}

TABS = [
    ("no_card",      "No Scorecard",     ["id", "name", "city", "st", "issue"]),
    ("no_rating",    "No Rated Tee",     ["id", "name", "city", "st", "issue"]),
    ("coverage",     "Unmapped & Partial",
     ["id", "name", "city", "st", "greens", "expected", "issue", "nines_missing",
      "legacy_base_greens", "note"]),
    ("par_mismatch", "Par Mismatches",
     ["id", "name", "city", "st", "agree", "reference", "source", "verified", "detail"]),
    ("shared_gps",   "Shared GPS",       ["ids", "names", "greens", "issue"]),
    ("wrong_course", "Wrong-Course Check",
     ["id", "name", "city", "st", "km", "greens", "source", "verified", "note"]),
    ("city",         "City Hygiene",     ["id", "name", "st", "lat", "lng", "issue"]),
    ("kevin",        "Kevin Queue",      ["id", "name", "city", "st", "status", "directive"]),
]

def write_xlsx(summary, findings, tn, out_path, previous=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    ws.append(["Bad Golf - Course Mapping Audit"]); ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Generated {summary['generated']} - live courses {summary['live_courses']:,} - "
               f"GPS rows {summary['gps_rows']:,} - wired three-nine facilities {summary['three_nine_facilities']}"])
    ws.append([])
    ws.append(["THREE-NINE AWARE. Wired three-nine facilities are exempt from the no-card and"])
    ws.append(["no-rated-tee checks: their pars and their per-combination ratings live in the app"])
    ws.append(["config, not in shared:courses. Their mapping is counted across per-nine rows"])
    ws.append(["(base#nine), never the base id. Do NOT 'fix' a three-nine by writing a card into"])
    ws.append(["shared:courses - that shadows the config and destroys the 27-hole data."])
    ws.append([])
    ws.append(["#", "Finding", "Count", "Prev", "Change", "Severity", "Detail tab", "Notes"])
    [setattr(c, "font", bold) for c in ws[ws.max_row]]
    for i, (key, tab, _cols) in enumerate(TABS, 1):
        n = len(findings.get(key, []))
        prev = (previous or {}).get(key)
        if prev is None:
            delta = ""
        elif n == prev:
            delta = "no change"
        elif n == 0:
            delta = "CLEARED (-%d)" % prev
        else:
            delta = ("%+d" % (n - prev))
        ws.append([i, tab, n, prev if prev is not None else "", delta,
                   SEVERITY.get(key, ""), tab, NOTES.get(key, "")])
    # ---- the numbers that actually move ----------------------------------
    # A facility short of ONE nine and a facility with nothing mapped both count
    # as exactly 1 row on line 3, so a day that recovers 981 greens can read "no
    # change". These lines measure the work itself, not the number of courses it
    # is spread across, and they are carried in the JSON so --prev tracks them.
    #
    # The two sign-off lines are NOT a backlog and must never be read as one.
    # Those facilities were signed off as 18-hole courses, correctly, with 18
    # greens and their fairway targets. Three-nine wiring was added afterwards
    # and raised the expectation to 27 without anyone asking the mapper for a
    # third nine. The sign-off stands; the extra nine is new work the wiring
    # created. Counting it against the sign-off is how you end up asking someone
    # to redo a job they already did right.
    m = summary.get("metrics") or {}
    pm = (previous or {}).get("_metrics") or {}
    if m:
        ws.append([])
        ws.append(["Work outstanding", "Now", "Prev", "Change", "", "", "", ""])
        [setattr(c, "font", bold) for c in ws[ws.max_row]]
        for label, key in (("Nines still to map", "nines_missing"),
                           ("Greens still to map", "greens_missing"),
                           ("Greens recovered by nine-split today", "greens_recovered"),
                           ("Signed off at 18, wiring later made it 27 (facilities)", "verified_short"),
                           ("  ... greens the wiring added after that sign-off", "verified_short_greens"),
                           ("Kevin queue = Incomplete chip", "reconciled")):
            if key not in m:
                continue
            p = pm.get(key)
            d = "" if p is None else ("no change" if p == m[key] else "%+d" % (m[key] - p))
            ws.append([label, m[key], p if p is not None else "", d])
    for col, w in zip("ABCDEFGH", (5, 22, 8, 8, 14, 10, 22, 96)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["A"].width = 42

    for key, tab, cols in TABS:
        s = wb.create_sheet(tab[:31])
        s.append(cols); [setattr(c, "font", bold) for c in s[1]]
        rows = findings.get(key, [])
        if not rows and key == "kevin" and findings.get("kevin_note"):
            s.append(["(not read)", str(findings["kevin_note"])])
        for r in (rows if isinstance(rows, list) else []):
            s.append([r.get(c, "") for c in cols])
        s.freeze_panes = "A2"
        for idx, c in enumerate(cols):
            s.column_dimensions[chr(65 + idx)].width = 46 if c in ("detail", "note", "names", "ids") else 20

    tabsheet = wb.create_sheet("Three-Nine Facilities")
    tabsheet.append(["course_id", "nines", "rated", "config source"])
    [setattr(c, "font", bold) for c in tabsheet[1]]
    for cid, cfg in sorted(tn.items()):
        tabsheet.append([cid, ", ".join(sorted(cfg["nines"])), "YES" if cfg["rated"] else "no",
                         cfg.get("source", "")])
    tabsheet.freeze_panes = "A2"
    for col, w in zip("ABCD", (48, 40, 8, 14)):
        tabsheet.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path


def load_previous(path):
    """Counts from a prior run, so the Summary can show movement rather than a
    snapshot. Accepts the .json this script writes alongside the xlsx."""
    if not path or not os.path.exists(path):
        return None
    try:
        prev = json.load(open(path)).get("summary", {})
        out = dict(prev.get("counts") or {})
        out["_metrics"] = prev.get("metrics") or {}
        return out
    except Exception:
        return None


def main():
    summary, findings, tn, out_path, args = build()
    previous = load_previous(args.prev)
    write_xlsx(summary, findings, tn, out_path, previous=previous)
    print(json.dumps(summary, indent=2))
    print("wrote", out_path)
    with open(out_path.replace(".xlsx", ".json"), "w") as f:
        json.dump({"summary": summary,
                   "findings": findings if args.json else {}}, f, indent=2)
    print("wrote", out_path.replace(".xlsx", ".json"), "(feed it to --prev next run)")


if __name__ == "__main__":
    main()
